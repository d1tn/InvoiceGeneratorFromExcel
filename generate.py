#モジュール読込
import openpyxl
from openpyxl.styles import Alignment, Font
import pandas as pd
import datetime
import calendar

##############################
#　　　　　　　初期設定
##############################

###################
# ユーザー基本設定
###################
# 法人名・屋号
corpName = '株式会社●●'

# 代表者氏名
topName = '山田　太郎'

# 郵便番号
postNum = '123-4567'

# 住所
Address = '東京都文京区◎◎1丁目1番地1号'

# 電話番号
telNum = "03-1234-5678"

# 振込先：銀行名、支店名
bankName = "●●●銀行 ▼▲支店"

# 振込先：口座情報
bankData = "普通 １２３　１２３４５６７　ヤマダ　タロウ"

# 税率
taxRate = 0.1

#保存先フォルダ
wbDir = '請求書/'


###################
# 出力する値とセル番地の指定
###################
#請求一覧の入力範囲
dataRange='A3:H13'

# 法人名・屋号
cell_corpName = 'G9'

# 代表者氏名
cell_topName = 'G10'

# 郵便番号
cell_postNum = 'G11'

# 住所
cell_Address = 'G12'

# 電話番号
cell_telNum =  'G13'

# 振込先：銀行名、支店名
cell_bankName = 'B31'

# 振込先：口座情報
cell_bankData = 'C32'

# 請求書日付
cell_invoiceDate = 'H1'
# 請求書No.
cell_invoiceNo = 'H2'
# 宛先（相手先）
cell_toName = 'A6'
# 受注No.　=>　.cell()で指定
cell_orderNo_row = 18
cell_orderNo_col = 1

# 合計請求額
cell_sumPrice = 'G28'
# お振込み期限
cell_dueDate = 'B35'

##############################
#　　　　　　　以下実装
##############################
# 入力値
acceptMonth = input('請求書を作成する年,月を半角数字6桁で入力してください。\n※空欄の場合は今月分の請求データを参照します\n　例：2021年9月 => 202109\n>>')

# 今日の年、月を取得
now_y = datetime.datetime.now().strftime('%Y')
now_m = datetime.datetime.now().strftime('%m')

# 入力値チェック(空白は現在の日付、数字以外・桁数違いはエラー)
if acceptMonth == '':
    acceptMonth = now_y + now_m
if not acceptMonth.isnumeric():
    print('半角数字のみ入力可能です。')
elif len(acceptMonth) != 6:
    print('西暦、月は6桁で入力してください。')
else:
    if int(acceptMonth[:4]) not in range(int(now_y)-30,int(now_y)+1):
        print('西暦を正しく入力してください。','\n※西暦は',int(now_y)-30,'~',int(now_y)+1,'の範囲で入力してください')
    elif int(acceptMonth[-2:]) not in range(1,13):
        print('月は01~12で入力してください。')
    else:
        try:
            # シート名の取得
            sheetName='請求一覧' + acceptMonth[2:]
            # ワークシートの読込
            wb = openpyxl.load_workbook('請求一覧.xlsx')
            ws = wb[sheetName]
        except KeyError:
            print('該当するシートが存在しません。')
        else:
            ws_customerTbl = wb['顧客管理テーブル']
            # 顧客CDのDF化
            customerTbl = []
            for row in ws_customerTbl['A2:B100']:
                values = []
                for col in row:
                    values.append(col.value)
                customerTbl.append(values)
            customerTbl = pd.DataFrame(customerTbl[1:], columns=customerTbl[0])

            # 請求一覧をDataFrame化
            df = []
            for row in ws[dataRange]:
                values = []
                for col in row:
                    values.append(col.value)
                df.append(values)
            df = pd.DataFrame(df[1:], columns=df[0])

            # 重複のない相手先コードリストを作成
            df_customer = df['相手先コード'].dropna().astype(int)
            customerList = []
            for i in df_customer:
                if i not in customerList:
                    customerList.append(i)

            # 相手先コードごとのループ処理
            for i, cusCode in enumerate(customerList):
                # 請求書No.の生成
                invoiceNo = acceptMonth[:2]+acceptMonth[-2:]+'{:0=2}'.format(i+1)

                #請求書テンプレートを開く
                wb_tmp = openpyxl.load_workbook('Invoice_template.xlsx')
                ws_tmp = wb_tmp.worksheets[0]
                ws_tmp.title = '請求書'+invoiceNo

                #合計、税額の初期化
                chargeSum = 0
                taxAmount = 0

                #ユーザーデータを出力
                ws_tmp[cell_corpName] = corpName
                ws_tmp[cell_topName] = '　  '+topName
                ws_tmp[cell_postNum] = '　  〒'+postNum
                ws_tmp[cell_Address] = '　  '+Address
                ws_tmp[cell_telNum] = '　  '+telNum
                ws_tmp[cell_bankName] = '振込先　：　'+bankName
                ws_tmp[cell_bankData] = bankData

                # 請求書No.の出力
                ws_tmp[cell_invoiceNo] = invoiceNo

                #相手先名を取得 => 出力
                customerName = customerTbl.loc[customerTbl['顧客ＣＤ'] == cusCode ]['顧客名'].item()
                ws_tmp[cell_toName] = customerName + "　御中"

                #検収日から請求日（当月末）、振込期限（翌月末）を取得 => 出力
                ws_tmp[cell_invoiceDate] = acceptMonth[:4]+'年'+acceptMonth[-2:]+'月'+str(calendar.monthrange(int(acceptMonth[:4]), int(acceptMonth[-2:]))[1])+'日'
                if int(acceptMonth[-2:]) == 12:
                       ws_tmp[cell_dueDate] = 'お振込み期限　：　' + str(int(acceptMonth[:4]) +1)+'年'+str(int(acceptMonth[-2:])-11)+'月'+str(calendar.monthrange(int(acceptMonth[:4]), int(acceptMonth[-2:]))[1])+'日'
                else:
                       ws_tmp[cell_dueDate] = 'お振込み期限　：　' + acceptMonth[:4]+'年'+str(int(acceptMonth[-2:])+1)+'月'+str(calendar.monthrange(int(acceptMonth[:4]), int(acceptMonth[-2:]))[1])+'日'

                #相手先コードに該当する行の取得
                customer = df[df['相手先コード']== cusCode ]

                # 該当する行ごとにループ処理
                # 行ごとに請求内容を出力
                for j, row in enumerate(customer.itertuples()):
                    cell_orderNo_row_tmp = 0
                    cell_orderNo_row_tmp += cell_orderNo_row
                    cell_orderNo_row_tmp += j*2

                    #受注No
                    ws_tmp.cell(row=cell_orderNo_row_tmp, column=cell_orderNo_col , value=row[1])

                    # 案件名１
                    ws_tmp.cell(row=cell_orderNo_row_tmp, column=cell_orderNo_col +1 , value=row[4])

                    # 案件名２が空欄の場合は中央寄せ、双方とも入力されている場合はそれぞれの行で文字寄せを行う
                    if row[5] is None:
                        ws_tmp.unmerge_cells(start_row=cell_orderNo_row_tmp, start_column=cell_orderNo_col+1, end_row=cell_orderNo_row_tmp, end_column = cell_orderNo_col+3 )
                        ws_tmp.unmerge_cells(start_row=cell_orderNo_row_tmp+1, start_column=cell_orderNo_col+1, end_row=cell_orderNo_row_tmp+1, end_column = cell_orderNo_col+3 )
                        ws_tmp.merge_cells(start_row=cell_orderNo_row_tmp, start_column=cell_orderNo_col+1, end_row=cell_orderNo_row_tmp+1, end_column = cell_orderNo_col+3 )
                    else:
                        ws_tmp.cell(row=cell_orderNo_row_tmp+1, column=cell_orderNo_col +1 , value=row[5])
                        ws_tmp.cell(row=cell_orderNo_row_tmp, column=cell_orderNo_col +1).alignment = Alignment( horizontal='left', vertical='bottom')
                        ws_tmp.cell(row=cell_orderNo_row_tmp+1, column=cell_orderNo_col +1 ).alignment = Alignment( horizontal='right', vertical='top')

                    # 数量
                    ws_tmp.cell(row=cell_orderNo_row_tmp, column=cell_orderNo_col +4, value=1)

                    # 単位
                    ws_tmp.cell(row=cell_orderNo_row_tmp, column=cell_orderNo_col +5 , value='式')

                    # 金額
                    ws_tmp.cell(row=cell_orderNo_row_tmp, column=cell_orderNo_col +6 , value=row[7])

                    # 備考
                    ws_tmp.cell(row=cell_orderNo_row_tmp, column=cell_orderNo_col +7 , value=row[8])
                    chargeSum += int(row[7])

                # 消費税
                taxAmount = int(chargeSum * taxRate)
                ws_tmp.cell(row=cell_orderNo_row_tmp+2, column=cell_orderNo_col +6 , value=taxAmount)

                #その他文字、書式
                ws_tmp.cell(row=cell_orderNo_row_tmp+3, column=cell_orderNo_col +1 , value='以上に掛かる消費税')
                ws_tmp.cell(row=cell_orderNo_row_tmp+4, column=cell_orderNo_col +1 , value='～　以　下　余　白　～')
                ws_tmp.cell(row=cell_orderNo_row_tmp+4, column=cell_orderNo_col +1).font= Font(bold=True)
                ws_tmp[cell_sumPrice] = chargeSum + taxAmount

                #名前を付けて保存
                wbName = wbDir + '御請求書_'+ acceptMonth[:4] + '年' + acceptMonth[-2:] + '月_No' + invoiceNo[-2:] + '【' + customerName + '様】.xlsx'
                try:
                    wb_tmp.save(wbName)
                    print('【OK】正常に保存されました。:' + wbName)
                except PermissionError:
                    print('【NG】請求書ファイルを閉じてください。:' + wbName)
